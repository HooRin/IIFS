import csv
import os
import random
import numpy as np
import pandas as pd
import torch
import openpyxl
from torch import cosine_similarity
from transformers import BertTokenizer, AutoModel
from datetime import datetime
os.environ['CUDA_VISIBLE_DEVICES'] = "1"
# 检查CUDA是否可用，并设置设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")
print(torch.__version__)
print(torch.cuda.is_available())
# 加载预训练的模型和分词器
tokenizer = BertTokenizer.from_pretrained('E:/LLM-model/bge-large-zh-v1.5-lzs')
model = AutoModel.from_pretrained('E:/LLM-model/bge-large-zh-v1.5-lzs')
model.to(device)  # 将模型移动到GPU
print("Model loading completed…………")
# Mean Pooling function with attention mask
def mean_pooling(model_output, attention_mask):
    token_embeddings = model_output.last_hidden_state
    input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
    return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)

def getToken(texts):
    # 将文本转化为BERT模型可识别的token序列
    encoded_texts = tokenizer(texts, return_tensors="pt", truncation=True, padding=True, max_length=512)
    encoded_texts = {key: val.to(device) for key, val in encoded_texts.items()}  # 将张量移动到GPU
    # 获取嵌入向量
    with torch.no_grad():
        model_output = model(**encoded_texts)
    # Perform mean pooling to get sentence embeddings
    sentence_embeddings = mean_pooling(model_output, encoded_texts['attention_mask'])

    return sentence_embeddings

def getBertSim(seedLaw, curLaw):
    similarities = cosine_similarity(seedLaw["embeddings"], curLaw["embeddings"])
    # 获取张量中的值
    value = similarities.item()

    return value


def hierarchical_clustering(lawList,save_excel,save_npy):
    # 初始化聚类集合，每个聚类存储完整数据对象和嵌入向量
    cluster_set = [[{
        "data": lawList[0],  # 存储完整数据对象
        "embedding": getToken(lawList[0]["QA_new"])  # 仅QA_new字段生成嵌入
    }]]

    similarity_threshold = 0.7
    t = 0

    for law_data in lawList[1:]:  # law_data是完整的原始数据对象
        t += 1
        if t % 1000 == 0:
            print(f"已处理：{t}\t{datetime.now().time()}")

        law_embedding = getToken(law_data["QA_new"])  # 仍使用QA_new生成嵌入
        found_cluster = False

        # 遍历现有聚类
        for cluster in cluster_set:
            # 提取该聚类所有嵌入向量
            embeddings = [item["embedding"] for item in cluster]
            stacked_embeddings = torch.stack(embeddings)
            cluster_center = torch.mean(stacked_embeddings, dim=0)

            # 计算相似度
            similarity = cosine_similarity(law_embedding, cluster_center)

            if similarity >= similarity_threshold:
                # 添加完整数据对象和嵌入
                cluster.append({
                    "data": law_data,
                    "embedding": law_embedding
                })
                found_cluster = True
                break

        if not found_cluster:
            cluster_set.append([{
                "data": law_data,
                "embedding": law_embedding
            }])

    # 构建包含所有字段的结果数据集
    all_json_data = []
    vec_x_v1 = []

    for cluster_id, cluster in enumerate(cluster_set):
        for item in cluster:
            # 复制原始数据的所有字段
            record = item["data"].copy()
            # 添加聚类标签
            record["label"] = cluster_id
            all_json_data.append(record)
            vec_x_v1.append(item["embedding"])

    # 保存完整数据
    df = pd.DataFrame(all_json_data)
    df.to_excel(save_excel, index=False)

    # 保存向量数据
    np.save(save_npy, vec_x_v1)

    return all_json_data
def loadData_xlsx(filename):
    all_json_data = []
    workbook = openpyxl.load_workbook(filename)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    flag = 0
    # 读取数据
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

        type = row[0].value
        id = row[1].value
        title = row[2].value
        section = row[3].value
        content  = row[4].value
        prompt = row[5].value
        QA = row[6].value
        query = row[7].value
        answer = row[8].value
        if type!="type":
            data_test = {
                "type": type,
                "id": id,
                "title": title,
                "section": section,
                "content": content,
                "prompt": prompt,
                "QA": QA,
                "query": query,
                "answer": answer
            }
            all_json_data.append(data_test)
    print(len(all_json_data))
    return all_json_data

def loadRawcorpus(all_json_data):
    instruction_list = []
    t = 0
    # 读取数据
    for i in all_json_data:
        type = i["type"]
        id = i["id"]
        title = i["title"]
        section = i["section"]
        content = i["content"]
        prompt = i["prompt"]
        QA = i["QA"]
        query = i["query"]
        answer = i["answer"]
        # instruction_list.append(d["instruction"])
        QA_new = "问题：" + query + "\n答案：" + answer
        data_set = {
            "type": type,
            "id": id,
            "title": title,
            "section": section,
            "content": content,
            "prompt": prompt,
            "QA": QA,
            "query": query,
            "answer": answer,
            "QA_new": QA_new
        }
        instruction_list.append(data_set)

    return instruction_list

if __name__ == '__main__':
    #train
    filename1 =  'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step1-1-QA_SFT_train.xlsx'
    save_excel = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step1-2-QA_SFT_train_claster.xlsx'
    save_npy =   'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step1-3-QA_SFT_train_claster.npy'
    all_json_data = loadData_xlsx(filename1)
    print(len(all_json_data))
    all_json_data = loadRawcorpus(all_json_data)
    seedLaw = hierarchical_clustering(all_json_data,save_excel,save_npy)  # 选择K个种子数据


    #valid
    filename1 =  'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step1-1-QA_SFT_valid.xlsx'
    save_excel = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step1-2-QA_SFT_valid_claster.xlsx'
    save_npy =   'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step1-3-QA_SFT_valid_claster.npy'
    all_json_data = loadData_xlsx(filename1)
    print(len(all_json_data))
    all_json_data = loadRawcorpus(all_json_data)
    seedLaw = hierarchical_clustering(all_json_data,save_excel,save_npy)  # 选择K个种子数据
