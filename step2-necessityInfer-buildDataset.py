import os

import random
import openpyxl
import json

import pandas as pd


def loadData(filename1,save_path):
    # 参数 定义
    t = 0
    all_json_data = []
    # 打开xlsx文件
    workbook = openpyxl.load_workbook(filename1)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        question = row[7].value
        answer = row[8].value
        data_train = {
            "instruction": question,
            "input": "",
            "output": answer,
            "system": "",
            "history": [[],[]],
        }
        data_train = dict(data_train)
        all_json_data.append(data_train)
    dump_json(all_json_data,save_path)


def dump_json(dataset, name):
    f = open(name + '.json', "w", encoding="utf-8")
    json.dump(dataset, f, ensure_ascii=False)
    f.close()

if __name__ == '__main__':
    #path = '/media/hjj/000C-C468/LLM/'
    current_dir = os.getcwd()  # 获取当前工作目录路径
    # 获取当前目录的上上一级目录路径
    path = os.path.abspath(os.path.dirname(current_dir))
    path = os.path.abspath(os.path.dirname(path))
    filename1 = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step1-2-QA_SFT_train_claster.xlsx'
    save_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step2-1-QA_SFT_train_necessityInfer'
    selected_sentences = loadData(filename1,save_path)
    # NERtoJson(selected_sentences, path)
    filename1 = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step1-2-QA_SFT_valid_claster.xlsx'
    save_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step2-1-QA_SFT_valid_necessityInfer'
    selected_sentences = loadData(filename1,save_path)