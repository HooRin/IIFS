import os
import openpyxl
import json

import pandas as pd


def loadData(filename_prediction,filename_label,save_path):

    all_json_data_label = []
    all_json_data_prediction = []
    all_json_data = []

    # 打开xlsx文件
    workbook = openpyxl.load_workbook(filename_prediction)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        text = row[0].value
        text = str(text)
        label_index = text.index("label")
        prediction_index = text.index("predict")
        answer = text[label_index+9:prediction_index-4]
        prediction = text[prediction_index+11:-2]
        dataset = {
            "necessity_answer": answer,
            "necessity_prediction": prediction,
        }
        all_json_data_prediction.append(dataset)

    # 打开xlsx文件
    workbook = openpyxl.load_workbook(filename_label)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        type = row[0].value
        id = row[1].value
        title = row[2].value
        section = row[3].value
        content = row[4].value
        prompt = row[5].value
        QA = row[6].value
        query = row[7].value
        answer = row[8].value
        cluster_QA = row[9].value
        cluster_label = row[10].value
        if type!="type":
            dataset = {
                "type": type,
                "id": id,
                "title": title,
                "section": section,
                "content": content,
                "prompt": prompt,
                "QA": QA,
                "query": query,
                "answer": answer,
                "cluster_QA": cluster_QA,
                "cluster_label": cluster_label,
            }
            all_json_data_label.append(dataset)
    print(len(all_json_data_label))

    # 同时遍历两个列表，合并字典
    for dict1, dict2 in zip(all_json_data_label, all_json_data_prediction):
        # 创建一个新的字典来存储合并后的结果
        merged_dict = dict1.copy()  # 复制第一个字典的内容
        merged_dict.update(dict2)  # 更新第二个字典的内容
        all_json_data.append(merged_dict)  # 将合并后的字典添加到结果列表中
        # 创建DataFrame
    df = pd.DataFrame(all_json_data)
    # 保存到Excel文件
    df.to_excel(save_path,index=False)
    df = pd.DataFrame(all_json_data_prediction)


if __name__ == '__main__':

    filename_label = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step1-2-QA_SFT_train_claster.xlsx'
    filename_prediction = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step2-2-QA_SFT_train_generated_predictions.xlsx'
    save_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step3-1-QA_SFT_train_claster_withPredictions.xlsx'
    selected_sentences = loadData(filename_prediction,filename_label,save_path)

    filename_label = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step1-2-QA_SFT_valid_claster.xlsx'
    filename_prediction = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step2-2-QA_SFT_valid_generated_predictions.xlsx'
    save_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step3-1-QA_SFT_valid_claster_withPredictions.xlsx'
    selected_sentences = loadData(filename_prediction,filename_label,save_path)
