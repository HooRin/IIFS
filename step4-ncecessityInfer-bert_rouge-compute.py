import json
import os
import sys
import openpyxl
import sys
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'
import pandas as pd
from openpyxl.workbook import Workbook
from bert_score import score
sys.path.append("E:/LLM-Project/PIXIU-main/src/financial-evaluation")
sys.path.append("E:/LLM-Project/PIXIU-main/src/evaluate-metric")
import jieba
from rouge_chinese import Rouge
from transformers import BertTokenizer, BertModel
os.environ['HF_HUB_OFFLINE'] = '1'

class QAwithString():
    def rougeChinese(self, items):

        jieba.load_userdict("E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/审计分词表.txt")
        hyps, refs = map(list, zip(*[[' '.join(jieba.cut(d[0])), ' '.join(jieba.cut(d[1]))] for d in items]))
        filter_hyps = []
        filter_refs = []
        for i in range(len(hyps)):
            hyp = hyps[i]
            ref = refs[i]
            # if self.is_whitespace_string(hyp) or self.is_whitespace_string(ref):
            #     continue
            if hyp != '' and ref != '':
                filter_hyps.append(hyp)
                filter_refs.append(ref)
        rouge = Rouge()
        scores = rouge.get_scores(filter_hyps, filter_refs, avg=False, ignore_empty=True)
        return scores


    def bert_score(self, items):
        golds, preds = zip(*items)
        # 加载本地的 bert-base-chinese 模型和分词器
        # fineTuningBert_audit
        # bert-base-chinese
        # tokenizer = BertTokenizer.from_pretrained("E:/LLM-model/bert-base-chinese")
        # model = BertModel.from_pretrained("E:/LLM-model/bert-base-chinese")

        # 使用 bert_score 计算 BERTScore
        # 注意：bert_score 的 score 函数需要模型和分词器作为参数
        P, R, F1 = score(golds, preds, lang="zh", model_type="bert-base-chinese", verbose=True)
        output_dict = {
            "precision": P.tolist(),
            "recall": R.tolist(),
            "f1": F1.tolist(),
        }
        return output_dict["f1"]

def compute_method(json_file_path,save_file_path):
    qa_task = QAwithString()

    # 加载原始数据
    workbook = openpyxl.load_workbook(json_file_path)
    sheet = workbook.worksheets[0]

    items = []
    original_data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        type = row[0].value
        id = row[1].value
        title = row[2].value
        section = row[3].value
        content = row[4].value
        prompt = row[5].value
        QA = row[6].value
        query = row[7].value
        answer = row[8].value
        cluster_QA = row[9].value
        cluster_label = row[10].value
        response = row[12].value
        if answer != "necessity_answer":
            my_tuple = (answer, response)
            items.append(my_tuple)
            dataset = {
                "type": type,
                "id": id,
                "title": title,
                "section": section,
                "content": content,
                "prompt": prompt,
                "QA": QA,
                "query": query,
                "answer": answer,
                "cluster_QA": cluster_QA,
                "cluster_label": cluster_label,
                "response": response,
            }
            original_data.append(dataset)

    print(f"Total items: {len(items)}")

    # 计算 ROUGE 分数
    rouge_chinese = qa_task.rougeChinese(items)

    # 每次处理的数据量
    chunk_size = 3000
    # 计算总共有多少个分块
    num_chunks = (len(items) + chunk_size - 1) // chunk_size
    # 存储所有分块的 BERT 分数
    all_bert_scores = []
    # 遍历每个分块
    for i in range(num_chunks):
        # 计算当前分块的起始和结束索引
        start = i * chunk_size
        end = min(start + chunk_size, len(items))

        # 获取当前分块的数据
        chunk = items[start:end]

        # 计算当前分块的 BERT 分数
        bert_score = qa_task.bert_score(chunk)
        for i in bert_score:
            # 将结果存储到列表中
            all_bert_scores.append(i)

        # 打印进度（可选）
        print(f"Processed chunk {i + 1}/{num_chunks}")

    print("总共计算了Bertscore数量：" + str(len(all_bert_scores)))
    print(all_bert_scores)
    result_save = []
    t = 0
    for index, (item1, item2, item3) in enumerate(zip(original_data, rouge_chinese, all_bert_scores)):
        dataset = {
            "type": item1["type"],
            "id": item1["id"],
            "title": item1["title"],
            "section": item1["section"],
            "content": item1["content"],
            "prompt": item1["prompt"],
            "QA": item1["QA"],
            "query": item1["query"],
            "answer": item1["answer"],
            "cluster_QA": item1["cluster_QA"],
            "cluster_label": item1["cluster_label"],
            "response": item1["response"],
            "rouge-l": item2["rouge-l"]['f'],
            "bertscore": item3,
        }
        t = t + 1
        result_save.append(dataset)

    df = pd.DataFrame(result_save)
    # 保存到Excel文件
    df.to_excel(save_file_path, index=False)


if __name__ == '__main__':
    json_file_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step3-1-QA_SFT_train_claster_withPredictions.xlsx"
    save_file_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step4-1-QA_SFT_train_Predictions_with_score.xlsx"
    compute_method(json_file_path, save_file_path)

    json_file_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step3-1-QA_SFT_valid_claster_withPredictions.xlsx"
    save_file_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step4-1-QA_SFT_valid_Predictions_with_score.xlsx"
    compute_method(json_file_path, save_file_path)





