import csv
import json
import os
import random
import numpy as np
import pandas as pd
import torch
import openpyxl
from torch import cosine_similarity
from transformers import BertTokenizer, AutoModel
from datetime import datetime
os.environ['CUDA_VISIBLE_DEVICES'] = "1"
# 检查CUDA是否可用，并设置设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")
print(torch.__version__)
print(torch.cuda.is_available())
# 加载预训练的模型和分词器
tokenizer = BertTokenizer.from_pretrained('E:/LLM-model/bge-large-zh-v1.5-lzs')
model = AutoModel.from_pretrained('E:/LLM-model/bge-large-zh-v1.5-lzs')
model.to(device)  # 将模型移动到GPU
print("Model loading completed…………")
# Mean Pooling function with attention mask
def mean_pooling(model_output, attention_mask):
    token_embeddings = model_output.last_hidden_state
    input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
    return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)

def getToken(texts):
    # 将文本转化为BERT模型可识别的token序列
    encoded_texts = tokenizer(texts, return_tensors="pt", truncation=True, padding=True, max_length=512)
    encoded_texts = {key: val.to(device) for key, val in encoded_texts.items()}  # 将张量移动到GPU
    # 获取嵌入向量
    with torch.no_grad():
        model_output = model(**encoded_texts)
    # Perform mean pooling to get sentence embeddings
    sentence_embeddings = mean_pooling(model_output, encoded_texts['attention_mask'])

    return sentence_embeddings

def getBertSim(seedLaw, curLaw):
    similarities = cosine_similarity(seedLaw["embeddings"], curLaw["embeddings"])
    # 获取张量中的值
    value = similarities.item()

    return value

def save_similarity_matrix_checkpoint(all_similarity_save, save_path):
    """将当前的相似度矩阵保存为 .npy 文件"""
    similarity_matrix = np.array(all_similarity_save)
    np.save(save_path, similarity_matrix)
    print(f"已保存相似度矩阵到 {save_path},共 "+str(len(similarity_matrix))+" 条")
    return similarity_matrix

def save_similarity_matrix_final(all_similarity_save, save_path):
    """将当前的相似度矩阵保存为 .npy 文件"""
    similarity_matrix = np.array(all_similarity_save)
    np.save(save_path, similarity_matrix)
    print(f"已保存相似度矩阵到 {save_path},共 "+str(len(similarity_matrix))+" 条")
    return similarity_matrix

def hierarchical_clustering(lawList,taskList):
    all_similarity_save = []
    all_similarity_checkpoint = []
    cluster_set = []  # task法规embedding保存
    save_threshold = 10000  # 设置保存阈值
    save_count = 0  # 保存计数器
    # 对每个task法规执行聚类分配
    for law in taskList:
        law_embedding = getToken(law)
        cluster_set.append(law_embedding)


    for t,i in enumerate(lawList):
        if t%1000==0:
            print("已跑："+str(t)+"\t" + str(datetime.now().time()))
        law_embedding = getToken(i)
        save_similarities = []

        #循环计算每一条的相似度
        for j in cluster_set:
            similarity = cosine_similarity(law_embedding, j)
            # 保留四位小数
            similarity = np.round(similarity.item(), 4)
            save_similarities.append(similarity)


        all_similarity_save.append(save_similarities)
        all_similarity_checkpoint.append(save_similarities)
            # 每处理 save_threshold 条数据，保存一次中间结果
        if (t + 1) % save_threshold == 0:
            save_count += 1


            save_path = f"E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train\data\MultiTaskCorrelation-checkpoint/train_similarity_chunk_{save_count}.npy"
            save_similarity_matrix_checkpoint(all_similarity_checkpoint, save_path)
            all_similarity_checkpoint = []  # 清空列表以便下一次保存

    # 保存剩余的数据（如果有的话）
    if all_similarity_checkpoint:
        save_count += 1
        save_path = f"E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train\data\MultiTaskCorrelation-checkpoint/train_similarity_chunk_{save_count}.npy"
        save_similarity_matrix_checkpoint(all_similarity_checkpoint, save_path)


    save_similarity_matrix_final(all_similarity_save,"E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train\data\MultiTaskCorrelation-checkpoint/train_similarity_matrix.npy")
    # 保存为 .npy 文件

def loadData_task(filename):
    all_json_data = []
    # 打开文件
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            # 解析每一行的JSON数据
            data = json.loads(line)
            # 打印或处理数据

    for i in data:
        if i["input"]==None:
            text = i['instruction'] +'\n'+i["output"]
        else:
            text = i['instruction'] +'\n'+i["input"]+'\n'+i["output"]
        all_json_data.append(text)
    print("加载任务"+str(filename)+"数据量：  "+str(len(all_json_data)))
    return all_json_data

def load_law(filename):
    # 加载原始数据
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.worksheets[0]

    items = []
    original_data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        type = row[0].value
        id = row[1].value
        title = row[2].value
        section = row[3].value
        content = row[4].value
        prompt = row[5].value
        QA = row[6].value
        query = row[7].value
        answer = row[8].value
        cluster_QA = row[9].value
        cluster_label = row[10].value
        response = row[11].value
        rougel = row[12].value
        bertscore = row[13].value

        items.append(cluster_QA)

        dataset = {
            "type":type,
            "id":id,
            "title":title,
            "section":section,
            "content":content,
            "prompt":prompt,
            "QA":QA,
            "query":query,
            "answer":answer,
            "clusterQA":cluster_QA,
            "cluster_label":cluster_label,
            "response":response,
            "rougel":rougel,
            "bertscore":bertscore,
        }
        original_data.append(dataset)

    return items,original_data

if __name__ == '__main__':
    #训练集
    task1_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/NER/train.jsonl"
    task2_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/处罚依据信息抽取/train.jsonl"
    task3_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/定性依据信息抽取/train.jsonl"
    task4_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/审计问题分析/train.jsonl"
    task5_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/审计问题定性/train.jsonl"
    task6_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/SFT data/data/审计问题总结/train.jsonl"

    all_json_data_qa = []
    selected_QA1 = loadData_task(task1_path)
    selected_QA2 = loadData_task(task2_path)
    selected_QA3 = loadData_task(task3_path)
    selected_QA4 = loadData_task(task4_path)
    selected_QA5 = loadData_task(task5_path)
    selected_QA6 = loadData_task(task6_path)

    for i in selected_QA1:
        all_json_data_qa.append(i)
    for i in selected_QA2:
        all_json_data_qa.append(i)
    for i in selected_QA3:
        all_json_data_qa.append(i)
    for i in selected_QA4:
        all_json_data_qa.append(i)
    for i in selected_QA5:
        all_json_data_qa.append(i)
    for i in selected_QA6:
        all_json_data_qa.append(i)

    print("*******************************************")
    print("总共："+str(len(all_json_data_qa)))

    # 加载法规数据
    law_path = "E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step4-1-QA_SFT_train_Predictions_with_score.xlsx"
    items,original_data = load_law(law_path)

    seedLaw = hierarchical_clustering(items, all_json_data_qa)  # 选择K个种子数据
    # NERtoJson(selected_sentences, path)

