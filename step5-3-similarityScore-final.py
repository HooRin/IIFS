import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
import pandas as pd


def similarityCompute(valid_similarity_path,law_path,save_path):
    loaded_similarity_matrix = np.load(valid_similarity_path[0])

    # 对每一行的元素进行排序（从高到低）
    sorted_indices = np.argsort(loaded_similarity_matrix, axis=1)[:, ::-1]
    sorted_similarity_matrix = np.take_along_axis(loaded_similarity_matrix, sorted_indices, axis=1)
    num_lists = sorted_similarity_matrix.shape[1]  # 获取列数

    # 指定需要绘制箱线图的列索引
    # valid:100
    # train:400
    target_ks = 400


    print("提取数据过程：")

    # 计算每一行前 k 个最大值的平均值 A
    top_k_averages = np.mean(sorted_similarity_matrix[:, :target_ks], axis=1)

    # 加载原始数据
    workbook = openpyxl.load_workbook(law_path)
    sheet = workbook.worksheets[0]
    original_data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        type = row[0].value
        id = row[1].value
        title = row[2].value
        section = row[3].value
        content = row[4].value
        prompt = row[5].value
        QA = row[6].value
        query = row[7].value
        answer = row[8].value
        cluster_QA = row[9].value
        cluster_label = row[10].value
        response = row[11].value
        rougel = row[12].value
        bertscore = row[13].value

        if type != "type":
            dataset = {
                "type": type,
                "id": id,
                "title": title,
                "section": section,
                "content": content,
                "prompt": prompt,
                "QA": QA,
                "query": query,
                "answer": answer,
                "cluster_QA": cluster_QA,
                "cluster_label": cluster_label,
                "response": response,
                "rougel":rougel,
                "bertscore":bertscore,
            }
            original_data.append(dataset)

    save_result = []
    # 同时遍历两个列表，合并字典
    for item1, item2 in zip(original_data, top_k_averages):
        dataset = {
            "type": item1["type"],
            "id": item1["id"],
            "title": item1["title"],
            "section": item1["section"],
            "content": item1["content"],
            "prompt": item1["prompt"],
            "QA": item1["QA"],
            "query": item1["query"],
            "answer": item1["answer"],
            "cluster_QA": item1["cluster_QA"],
            "cluster_label": item1["cluster_label"],
            "response": item1["response"],
            "rouge-l": item1["rougel"],
            "bertScore": item1["bertscore"],
            "relevanceScore":item2
        }
        save_result.append(dataset)


    # 创建DataFrame
    df = pd.DataFrame(save_result)
    # 保存到Excel文件
    df.to_excel(save_path,index=False)


if __name__ == '__main__':
    valid_similarity_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/MultiTaskCorrelation-checkpoint/train_similarity_matrix.npy',
    law_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step4-1-QA_SFT_train_Predictions_with_score.xlsx'
    save_path = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step5-1-QA_SFT_relevanceScore.xlsx'

    similarityCompute(valid_similarity_path,law_path,save_path)