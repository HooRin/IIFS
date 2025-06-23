import math
from collections import defaultdict

import numpy as np
import openpyxl
import pandas as pd



def loadData(filename_evaluate,save_all,save_QA):

    all_json_data = []

    # 打开xlsx文件
    workbook = openpyxl.load_workbook(filename_evaluate)
    # 获取第一个工作表
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        type = row[0].value

        if type!="type":
            dataset = {
                'type':row[0].value,
                'id': row[1].value,
                'title': row[2].value,
                'section': row[3].value,
                'content': row[4].value,
                'prompt': row[5].value,
                'QA': row[6].value,
                'query': row[7].value,
                'answer': row[8].value,
                'cluster_QA': row[9].value,
                'cluster_label': row[10].value,
                'response': row[11].value,
                'rougel': float(row[12].value),
                'bertScore': float(row[13].value),
                'relevanceScore': float(row[14].value),
                'ifSelect':1,
                'comprehensiveScore':0
            }
            all_json_data.append(dataset)


    print("数据个数："+str(len(all_json_data)))
    # 计算中位数
    mid_rougel = []
    mid_bertScore = []
    mid_relevanceScore = []
    for i in all_json_data:
        mid_rougel.append(i['rougel'])
        mid_bertScore.append(i['bertScore'])
        mid_relevanceScore.append(i['relevanceScore'])
    median_rougel = np.median(mid_rougel)
    median_bertScore = np.median(mid_bertScore)
    median_relevanceScore = np.median(mid_relevanceScore)
    print("rougel_median:" + str(median_rougel))
    print("bertScore_median:" + str(median_bertScore))
    print("relevanceScore_median:" + str(median_relevanceScore))


    t= 0
    for i in all_json_data:
        # 加权留作排名使用
        # 2.0-1 冗余性、必要性
        # i['Score_with_rouge_bert'] = 0.7 * i['text_rouge_l_f'] + 0.3 * i['text_bert']
        # if i['text_rouge_l_f'] > median_rougel and i['text_bert'] > median_bertScore:
        #     i['ifSelect'] = 0
        #     t = t + 1
        # 2.0-2 冗余性、必要性、任务相关性
        i['comprehensiveScore'] = 0.3*(1-(0.7*i['rougel']+0.3*i['bertScore']))+0.7*i['relevanceScore']

        if i['rougel'] > median_rougel and i['bertScore'] > median_bertScore and i['relevanceScore'] < median_relevanceScore:
            i['ifSelect'] = 0
            t = t + 1
    print("过滤："+str(t))
    # 假设 all_json_data 是你已经加载和处理过的数据列表
    grouped_data = defaultdict(list)

    # 将数据按 text_label 分组
    for item in all_json_data:
        if item['ifSelect']==1:
            grouped_data[item['cluster_label']].append(item)

    # 对每个分组内的数据按 Score_with_rouge_bert 升序排序
    for label in grouped_data:
        grouped_data[label].sort(key=lambda x: x['comprehensiveScore'], reverse=True)
    # 获取调优数据
    bad_data = {}
    sum = 0
    for label, items in grouped_data.items():
        num_data = len(items)
        if num_data <= 5:
            num_select = 1
        else:
            num_select = 2 ** math.ceil(math.log(num_data/5, 2))

        # 保留排名靠前的数据
        bad_data[label] = items[:num_select]


    grouped_data = bad_data
    # 扁平化数据
    flat_data = []
    for label, items in grouped_data.items():
        for item in items:
            item['cluster_label'] = label  # 将标签添加回每个数据项
            flat_data.append(item)

    # 创建DataFrame
    df = pd.DataFrame(flat_data)
    # 保存到Excel文件
    df.to_excel(save_all, index=False)
    print("筛选："+str(len(flat_data)))
    all_json_data = []
    for i in flat_data:
        dataset={
            "qaestion":i['query'],
            "answer":i['answer'],
        }
        dataset =dict(dataset)
        all_json_data.append(dataset)
    # 创建DataFrame
    df = pd.DataFrame(all_json_data)
    # 保存到Excel文件
    df.to_excel(save_QA, index=False)


if __name__ == '__main__':
    print("*************************************")
    print("***           筛选 valid           ***")
    print("*************************************")
    filename_evaluate = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step5-1-QA_SFT_relevanceScore.xlsx'
    save_all = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step6-1-QA_SFT_final_All.xlsx'
    save_QA = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/valid/data/step6-1-QA_SFT_valid_final_QA.xlsx'
    print(filename_evaluate)
    selected_sentences = loadData(filename_evaluate,save_all,save_QA)

    print("*************************************")
    print("***           筛选 train           ***")
    print("*************************************")
    filename_evaluate = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step5-1-QA_SFT_relevanceScore.xlsx'
    save_all = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step6-1-QA_SFT_final_All.xlsx'
    save_QA = 'E:/LLM-Project/算法设计/IFT-Select-necessarily（2025-3-13）/Singal pass-based IFT/train/data/step6-1-QA_SFT_train_final_QA.xlsx'
    print(filename_evaluate)
    selected_sentences = loadData(filename_evaluate,save_all,save_QA)